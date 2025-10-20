import openpyxl

wb = openpyxl.load_workbook(filename='book1.xlsx')

ws = wb.active

ws.cell(row=7,column=1,value='king')
ws.cell(row=7,column=2,value=1)

ws.append(['Devil',666])

wb.save('book1.xlsx')
wb.close()
