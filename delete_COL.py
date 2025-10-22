import openpyxl

wb = openpyxl.load_workbook(filename='new.xlsx')

ws = wb.active

delele_col = 2

print("before:",ws.max_column)

ws.delete_cols(delele_col)

print("after:", ws.max_column)

wb.save('new.xlsx')
wb.close()