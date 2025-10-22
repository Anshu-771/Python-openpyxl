import openpyxl
from openpyxl.styles import Font,Alignment

# print(dir(Font)) # task using Font

wb = openpyxl.load_workbook(filename='marks.xlsx')

ws = wb.active

for col in ws.iter_rows(min_row=1,max_row=1):   # BOLD
    for cell in col:
        cell.font = Font(bold=True)

for row in ws.iter_cols(min_col=1,max_col=1,min_row=2):  # ITALIC
    for cell in row:
        cell.font = Font(italic=True)

for CD in ws.iter_rows(min_row=2,max_row=4,min_col=2,max_col=3): # UNDERLINE
    for cell in CD:
        if(cell.value < 70):
            cell.font = Font(underline='single',color='ff0000')

for col in ws.iter_rows(min_row=1,max_row=1):   # SIZE,CENTER,FONT
    for cell in col:
        cell.font = Font(size=34,name='bodoni mt black')
        cell.alignment = Alignment(horizontal='center')

for col in ws.columns:
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    ws.column_dimensions[col[0].column_letter].width = max_length + 2



wb.save('marks.xlsx')
wb.close()