import openpyxl

wb = openpyxl.load_workbook(filename='book2.xlsx')

ws1 = wb['Sheet1']
ws2 = wb['Sheet2']
ws3 = wb['Sheet3']

# cell 1 from sheet 1

c1_ws1 = ws1.cell(row=1,column=1).value
print(c1_ws1)

# cell 1 from sheet 2

c1_ws2 = ws2.cell(row=1,column=1).value
print(c1_ws2)

# cell 1 from sheet 2

c1_ws3 = ws3.cell(row=1,column=1).value
print(c1_ws3)

wb.close()