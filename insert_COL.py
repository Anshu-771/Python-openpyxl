import openpyxl

wb = openpyxl.load_workbook(filename='new.xlsx')

ws = wb.active

insert_new_col = 1

ws.insert_cols(insert_new_col)

ws['A1'] = "hello again"
ws['A2'] = 'dupilicate'

wb.save('new.xlsx')

wb.close()
