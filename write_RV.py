import openpyxl

wb = openpyxl.load_workbook(filename='book2.xlsx')

ws = wb.active

ws['A1'and 'A2'] = ''
ws['A1'] = 'i am sheet 1'

wb.save('book2.xlsx')

wb.close()