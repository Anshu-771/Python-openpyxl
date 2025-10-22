import openpyxl

wb = openpyxl.load_workbook(filename='book2.xlsx')

ws1 = wb.worksheets[0]
ws2 = wb.worksheets[1]
ws3 = wb.worksheets[2]

ws1.cell(row=3,column=3,value='sheet 1 this!')
ws2.cell(row=3,column=3,value='sheet 2 this!')
ws3['C3'] = 'sheet 3 this!'

wb.save('book2.xlsx')
wb.close()