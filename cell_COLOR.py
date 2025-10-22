import openpyxl
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook(filename='new.xlsx')

ws = wb.active


ws['A1'].fill = PatternFill(start_color='ff0000',end_color='ff0000',fill_type='solid')

wb.save('new.xlsx')
wb.close()