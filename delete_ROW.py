import openpyxl

wb = openpyxl.load_workbook(filename='book1.xlsx')

ws = wb.active

print("before :" , ws.max_row)

delete_row = 10

ws.delete_rows(delete_row)

print("after : " , ws.max_row)

wb.save('book1.xlsx')

wb.close()